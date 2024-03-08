from selenium.webdriver.common.by import By
# Explicit Wait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.reader.excel import load_workbook

class WebMethods:


    def __init__(self):
        self.fileName = "Data/DDTF Frame Work.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.fileName)
        self.sheet = self.workbook[self.sheetName]

    def rowCount(self):
        """
        This method returns the maximum number of rows present in the Sheet 1
        :return: int
        """
        return self.sheet.max_row

    def readData(self, row, column):
        """
        This method will return the data present in the particular cell in the Sheet 1
        :param row:
        :param column:

        """
        return self.sheet.cell(row, column).value

    def writeData(self, row, column, data):
        """
        This method allows us to write the result in the Excel file
        :param row:
        :param column:
        :param data:
        """
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.fileName)


    def enterText(self, driver, locator, textValue):
        """
        This method find the element in web page and send text to input field
        :param driver:
        :param locator:
        :param textValue:
        """
        self.wait = WebDriverWait(driver, 10)
        element=self.wait.until(EC.presence_of_element_located((By.XPATH, locator)))
        element.clear()
        element.send_keys(textValue)
    def clickButton(self, driver, locator):
        """
        This Method find the element  in web page and click the button
        :param driver:
        :param locator:
        """
        self.wait = WebDriverWait(driver, 10)
        self.wait.until(EC.presence_of_element_located((By.XPATH, locator))).click()
